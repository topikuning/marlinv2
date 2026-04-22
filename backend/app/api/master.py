import io
import os
import tempfile
from fastapi import APIRouter, Depends, HTTPException, Request, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_
from typing import Optional, List
from datetime import datetime

from app.core.database import get_db
from app.models.models import Company, PPK, MasterWorkCode, User, WorkCategory
from app.schemas.schemas import (
    CompanyCreate, CompanyUpdate, CompanyOut,
    PPKCreate, PPKUpdate, PPKOut,
    MasterWorkCodeCreate, MasterWorkCodeOut,
    ExcelImportResult,
)
from app.api.deps import get_current_user, require_permission
from app.services.audit_service import log_audit

router = APIRouter(prefix="/master", tags=["master"])


# ═══════════════════════════════════════════ COMPANIES ═══════════════════════

@router.get("/companies", response_model=dict)
def list_companies(
    q: Optional[str] = None,
    is_active: Optional[bool] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    query = db.query(Company).filter(Company.deleted_at.is_(None))
    if q:
        query = query.filter(or_(Company.name.ilike(f"%{q}%"), Company.npwp.ilike(f"%{q}%")))
    if is_active is not None:
        query = query.filter(Company.is_active == is_active)
    total = query.count()
    items = query.order_by(Company.name).offset((page - 1) * page_size).limit(page_size).all()
    return {"total": total, "page": page, "page_size": page_size,
            "items": [CompanyOut.model_validate(i).model_dump(mode="json") for i in items]}


@router.post("/companies", response_model=dict)
def create_company(
    data: CompanyCreate, request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("master.create")),
):
    """
    Create a Company and — for type=contractor|consultant — auto-provision
    a User account with role "kontraktor" or "konsultan" respectively.
    The user starts with password "Ganti@123!" and must_change_password=True.
    Supplier companies do NOT get a user.
    """
    payload = data.model_dump()
    # company_type is required for auto-provisioning; default to contractor
    # if the client didn't pass it.
    if not payload.get("company_type"):
        payload["company_type"] = "contractor"

    c = Company(**payload)
    db.add(c)
    db.flush()  # need c.id for provisioning

    from app.services.user_provisioning_service import provision_user_for_company
    user, created = provision_user_for_company(db, c, created_by_id=current_user.id)

    db.commit()
    db.refresh(c)

    audit_payload = {"company_type": c.company_type}
    if user and created:
        audit_payload["auto_provisioned_user_id"] = str(user.id)
        audit_payload["auto_provisioned_username"] = user.username
    log_audit(db, current_user, "create", "company", str(c.id),
              changes=audit_payload, request=request, commit=True)

    return {
        "id": str(c.id),
        "success": True,
        "auto_provisioned_user": {
            "id": str(user.id),
            "username": user.username,
            "email": user.email,
            "default_password": "Ganti@123!",
            "must_change_password": True,
        } if (user and created) else None,
    }


@router.put("/companies/{company_id}", response_model=dict)
def update_company(
    company_id: str, data: CompanyUpdate, request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("master.update")),
):
    c = db.query(Company).filter(Company.id == company_id, Company.deleted_at.is_(None)).first()
    if not c:
        raise HTTPException(404, "Perusahaan tidak ditemukan")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(c, k, v)
    db.commit()
    log_audit(db, current_user, "update", "company", str(c.id), request=request, commit=True)
    return {"success": True}


@router.delete("/companies/{company_id}", response_model=dict)
def delete_company(
    company_id: str, request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("master.delete")),
):
    c = db.query(Company).filter(Company.id == company_id, Company.deleted_at.is_(None)).first()
    if not c:
        raise HTTPException(404, "Perusahaan tidak ditemukan")
    c.deleted_at = datetime.utcnow()
    c.is_active = False
    db.commit()
    log_audit(db, current_user, "delete", "company", str(c.id), request=request, commit=True)
    return {"success": True}


# ═══════════════════════════════════════════ PPK ═════════════════════════════

@router.get("/ppk", response_model=dict)
def list_ppk(
    q: Optional[str] = None,
    is_active: Optional[bool] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    query = db.query(PPK).filter(PPK.deleted_at.is_(None))
    if q:
        query = query.filter(or_(PPK.name.ilike(f"%{q}%"), PPK.nip.ilike(f"%{q}%"),
                                 PPK.satker.ilike(f"%{q}%")))
    if is_active is not None:
        query = query.filter(PPK.is_active == is_active)
    total = query.count()
    items = query.order_by(PPK.name).offset((page - 1) * page_size).limit(page_size).all()
    return {"total": total, "page": page, "page_size": page_size,
            "items": [PPKOut.model_validate(i).model_dump(mode="json") for i in items]}


@router.post("/ppk", response_model=dict)
def create_ppk(
    data: PPKCreate, request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("master.create")),
):
    """
    Create a PPK and auto-provision a User with role "ppk".
    Password "Ganti@123!" + must_change_password=True.
    """
    p = PPK(**data.model_dump())
    db.add(p)
    db.flush()

    from app.services.user_provisioning_service import provision_user_for_ppk
    user, created = provision_user_for_ppk(db, p, created_by_id=current_user.id)

    db.commit()
    db.refresh(p)

    log_audit(
        db, current_user, "create", "ppk", str(p.id),
        changes={"auto_provisioned_user_id": str(user.id) if created else None,
                 "username": user.username if created else None},
        request=request, commit=True,
    )
    return {
        "id": str(p.id),
        "success": True,
        "auto_provisioned_user": {
            "id": str(user.id),
            "username": user.username,
            "email": user.email,
            "default_password": "Ganti@123!",
            "must_change_password": True,
        } if created else None,
    }


@router.put("/ppk/{ppk_id}", response_model=dict)
def update_ppk(
    ppk_id: str, data: PPKUpdate, request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("master.update")),
):
    p = db.query(PPK).filter(PPK.id == ppk_id, PPK.deleted_at.is_(None)).first()
    if not p:
        raise HTTPException(404, "PPK tidak ditemukan")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(p, k, v)
    db.commit()
    log_audit(db, current_user, "update", "ppk", str(p.id), request=request, commit=True)
    return {"success": True}


@router.delete("/ppk/{ppk_id}", response_model=dict)
def delete_ppk(
    ppk_id: str, request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("master.delete")),
):
    p = db.query(PPK).filter(PPK.id == ppk_id, PPK.deleted_at.is_(None)).first()
    if not p:
        raise HTTPException(404, "PPK tidak ditemukan")
    p.deleted_at = datetime.utcnow()
    p.is_active = False
    db.commit()
    log_audit(db, current_user, "delete", "ppk", str(p.id), request=request, commit=True)
    return {"success": True}


# ═══════════════════════════════════════════ MASTER WORK CODES ═══════════════

@router.get("/work-codes", response_model=List[MasterWorkCodeOut])
def list_work_codes(
    q: Optional[str] = None,
    category: Optional[str] = None,
    page_size: int = Query(50, ge=1, le=1000),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """
    List master work codes. Supports fuzzy search via `q` (searches code,
    description, keywords, sub_category) so the BOQ item picker can show
    relevant results as the user types.
    """
    query = db.query(MasterWorkCode).filter(MasterWorkCode.is_active == True)  # noqa: E712
    if category:
        query = query.filter(MasterWorkCode.category == category)
    if q:
        term = f"%{q}%"
        from sqlalchemy import or_
        query = query.filter(or_(
            MasterWorkCode.code.ilike(term),
            MasterWorkCode.description.ilike(term),
            MasterWorkCode.keywords.ilike(term),
            MasterWorkCode.sub_category.ilike(term),
        ))
    return (
        query
        .order_by(MasterWorkCode.category, MasterWorkCode.code)
        .limit(page_size)
        .all()
    )


@router.post("/work-codes", response_model=dict)
def create_work_code(
    data: MasterWorkCodeCreate, request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("master.create")),
):
    if db.query(MasterWorkCode).filter(MasterWorkCode.code == data.code).first():
        raise HTTPException(400, "Kode sudah ada")
    m = MasterWorkCode(**data.model_dump())
    db.add(m)
    db.commit()
    log_audit(db, current_user, "create", "master_work_code", data.code, request=request, commit=True)
    return {"code": data.code, "success": True}


@router.put("/work-codes/{code}", response_model=dict)
def update_work_code(
    code: str, data: dict, request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("master.update")),
):
    m = db.query(MasterWorkCode).filter(MasterWorkCode.code == code).first()
    if not m:
        raise HTTPException(404, "Kode tidak ditemukan")
    for k, v in data.items():
        if hasattr(m, k) and k != "code":
            setattr(m, k, v)
    db.commit()
    log_audit(db, current_user, "update", "master_work_code", code, request=request, commit=True)
    return {"success": True}


@router.delete("/work-codes/{code}", response_model=dict)
def delete_work_code(
    code: str, request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("master.delete")),
):
    m = db.query(MasterWorkCode).filter(MasterWorkCode.code == code).first()
    if not m:
        raise HTTPException(404, "Kode tidak ditemukan")
    m.is_active = False
    db.commit()
    log_audit(db, current_user, "delete", "master_work_code", code, request=request, commit=True)
    return {"success": True}


# ═══════════════════════════════════════════ TEMPLATE + BULK IMPORT ══════════

_WORK_CODE_COLUMNS = [
    ("code",         True,  "Kode unik (mis. PER-001). Huruf kapital + angka/dash."),
    ("category",     True,  "persiapan | struktural | arsitektural | mep | site_work | khusus"),
    ("sub_category", False, "Label sub-kategori opsional (mis. Pondasi, Dinding)."),
    ("description",  True,  "Uraian pekerjaan — yang muncul di picker BOQ."),
    ("default_unit", False, "Satuan umum (mis. m2, m3, kg, ls)."),
    ("keywords",     False, "Kata kunci tambahan untuk pencarian, dipisah koma."),
    ("notes",        False, "Catatan internal opsional."),
]


def _build_work_code_template_bytes() -> bytes:
    """
    Bangun template .xlsx dua-sheet untuk import Master Kode Pekerjaan.
    Sheet 1 "Kode" — header + 3 baris contoh siap ditimpa.
    Sheet 2 "Petunjuk" — penjelasan per-kolom + daftar kategori valid.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Kode"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E40AF")
    for idx, (name, required, _) in enumerate(_WORK_CODE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=name + (" *" if required else ""))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(14, len(name) + 6)

    # 3 baris contoh dari 3 kategori berbeda supaya user tahu bentuk data.
    examples = [
        ("PER-001", "persiapan", "Mobilisasi", "Mobilisasi alat berat dan personil",
         "ls", "mob, demob", "Termasuk demobilisasi"),
        ("STR-B03", "struktural", "Beton", "Pengecoran kolom beton bertulang K-300",
         "m3", "kolom, beton, K-300", ""),
        ("MEP-L02", "mep", "Listrik", "Instalasi kabel NYY 4x10mm dalam conduit",
         "m", "kabel, listrik, nyy", ""),
    ]
    for r_idx, row in enumerate(examples, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    ws.freeze_panes = "A2"

    # Sheet petunjuk
    guide = wb.create_sheet("Petunjuk")
    guide.column_dimensions["A"].width = 20
    guide.column_dimensions["B"].width = 12
    guide.column_dimensions["C"].width = 80
    guide.append(["Kolom", "Wajib?", "Keterangan"])
    for cell in guide[1]:
        cell.font = header_font
        cell.fill = header_fill
    for name, required, desc in _WORK_CODE_COLUMNS:
        guide.append([name, "Ya" if required else "—", desc])

    guide.append([])
    guide.append(["Aturan umum:"])
    for rule in [
        "• Baris kosong diabaikan.",
        "• Kolom header tidak boleh diubah namanya (huruf kecil, persis seperti di atas).",
        "• Kalau kode sudah ada, baris akan di-skip (bukan update) — hapus dulu kalau ingin mengganti.",
        "• Kategori harus salah satu dari nilai enum yang tertera; huruf besar/kecil tidak masalah.",
    ]:
        guide.append([rule])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/work-codes/template")
def download_work_code_template(
    _: User = Depends(get_current_user),
):
    """
    Download template xlsx dua-sheet untuk mengisi Master Kode Pekerjaan
    secara massal. Pakai endpoint import-excel untuk meng-upload kembali.
    """
    data = _build_work_code_template_bytes()
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="template_master_kode_pekerjaan.xlsx"',
        },
    )


@router.post("/work-codes/import-excel", response_model=ExcelImportResult)
async def import_work_codes_excel(
    file: UploadFile = File(...),
    request: Request = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("master.create")),
):
    """
    Import massal Master Kode Pekerjaan dari file xlsx hasil template.

    Perilaku:
      - Kode yang sudah ada di DB → skip (dihitung di items_skipped),
        agar pengguna tidak kaget datanya tertimpa.
      - Baris dengan field wajib kosong (code, category, description) →
        skip + dicatat di errors[] supaya user tahu baris mana yang
        dilewati dan kenapa.
      - Kategori dinormalisasi ke lowercase dan divalidasi terhadap enum
        WorkCategory. Kategori tidak valid → skip + error.
    """
    import pandas as pd

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    result = ExcelImportResult(success=False)
    valid_categories = {c.value for c in WorkCategory}

    try:
        # Sheet pertama diambil, header = row 1. Pakai dtype=object agar
        # angka yang kebetulan dipakai sebagai kode (jarang, tapi mungkin)
        # tidak berubah jadi float.
        df = pd.read_excel(tmp_path, sheet_name=0, dtype=object)
        cols = [str(c).strip().lower().rstrip(" *") for c in df.columns]
        required = {"code", "category", "description"}
        missing = required - set(cols)
        if missing:
            result.errors.append(
                f"Kolom wajib tidak ditemukan: {', '.join(sorted(missing))}. "
                f"Pakai template dari tombol Template."
            )
            return result

        created_codes: list[str] = []
        for idx, row in df.iterrows():
            rec = {cols[i]: row.iloc[i] for i in range(len(cols))}

            def _s(key):
                v = rec.get(key)
                if v is None:
                    return ""
                s = str(v).strip()
                return "" if s.lower() == "nan" else s

            code = _s("code").upper()
            raw_cat = _s("category").lower()
            description = _s("description")

            if not code and not description:
                # Baris benar-benar kosong, diam-diam skip tanpa error.
                continue
            if not code:
                result.items_skipped += 1
                result.errors.append(f"Baris {idx + 2}: code kosong — skip.")
                continue
            if not description:
                result.items_skipped += 1
                result.errors.append(f"Baris {idx + 2}: description kosong — skip.")
                continue
            if raw_cat not in valid_categories:
                result.items_skipped += 1
                result.errors.append(
                    f"Baris {idx + 2}: kategori '{raw_cat}' tidak valid. "
                    f"Gunakan salah satu: {', '.join(sorted(valid_categories))}."
                )
                continue

            if db.query(MasterWorkCode).filter(MasterWorkCode.code == code).first():
                result.items_skipped += 1
                continue

            db.add(MasterWorkCode(
                code=code,
                category=WorkCategory(raw_cat),
                sub_category=_s("sub_category") or None,
                description=description,
                default_unit=_s("default_unit") or None,
                keywords=_s("keywords") or None,
                notes=_s("notes") or None,
            ))
            created_codes.append(code)
            result.items_imported += 1

        if created_codes:
            db.commit()
            log_audit(
                db, current_user, "bulk_create", "master_work_code",
                entity_id=None,
                changes={"imported_count": len(created_codes), "codes": created_codes[:50]},
                request=request, commit=True,
            )
        result.success = True
    except Exception as e:
        result.errors.append(f"Gagal membaca file: {e}")
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    return result


# ═══════════════════════════════════════════ MASTER FACILITIES ═══════════════
# Catalog of standard facility types (Gudang Beku, Pabrik Es, Cool Box, etc.)
# When creating a Facility under a Location the admin must pick from this list
# instead of typing free text. That keeps BOQ mapping consistent across sites
# and lets us seed the catalog from the BOQ Excel you provided.

from app.models.models import MasterFacility


@router.get("/facilities", response_model=dict)
def list_master_facilities(
    q: Optional[str] = None,
    is_active: Optional[bool] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    query = db.query(MasterFacility)
    if q:
        query = query.filter(or_(
            MasterFacility.name.ilike(f"%{q}%"),
            MasterFacility.code.ilike(f"%{q}%"),
        ))
    if is_active is not None:
        query = query.filter(MasterFacility.is_active == is_active)

    total = query.count()
    items = (
        query.order_by(MasterFacility.display_order, MasterFacility.name)
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return {
        "total": total, "page": page, "page_size": page_size,
        "items": [
            {
                "id": str(i.id),
                "code": i.code,
                "name": i.name,
                "facility_type": i.facility_type,
                "typical_unit": i.typical_unit,
                "description": i.description,
                "display_order": i.display_order,
                "is_active": i.is_active,
            }
            for i in items
        ],
    }


@router.post("/facilities", response_model=dict)
def create_master_facility(
    payload: dict, request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("master.create")),
):
    code = (payload.get("code") or "").strip().upper()
    name = (payload.get("name") or "").strip()
    if not code or not name:
        raise HTTPException(400, "Field 'code' dan 'name' wajib diisi.")
    if db.query(MasterFacility).filter(MasterFacility.code == code).first():
        raise HTTPException(400, f"Kode master fasilitas '{code}' sudah ada.")
    m = MasterFacility(
        code=code,
        name=name,
        facility_type=payload.get("facility_type") or "perikanan",
        typical_unit=payload.get("typical_unit"),
        description=payload.get("description"),
        display_order=int(payload.get("display_order") or 0),
        is_active=bool(payload.get("is_active", True)),
    )
    db.add(m)
    db.commit()
    db.refresh(m)
    log_audit(db, current_user, "create", "master_facility", str(m.id),
              changes={"code": code}, request=request, commit=True)
    return {"id": str(m.id), "success": True}


@router.put("/facilities/{facility_id}", response_model=dict)
def update_master_facility(
    facility_id: str, payload: dict, request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("master.update")),
):
    m = db.query(MasterFacility).filter(MasterFacility.id == facility_id).first()
    if not m:
        raise HTTPException(404, "Master fasilitas tidak ditemukan")
    for k in ("name", "facility_type", "typical_unit", "description",
              "display_order", "is_active"):
        if k in payload:
            setattr(m, k, payload[k])
    db.commit()
    log_audit(db, current_user, "update", "master_facility", str(m.id),
              request=request, commit=True)
    return {"success": True}


@router.delete("/facilities/{facility_id}", response_model=dict)
def delete_master_facility(
    facility_id: str, request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("master.delete")),
):
    """Soft delete via is_active=False. We never hard-delete because
    existing Facility rows may FK to this record."""
    m = db.query(MasterFacility).filter(MasterFacility.id == facility_id).first()
    if not m:
        raise HTTPException(404, "Master fasilitas tidak ditemukan")
    m.is_active = False
    db.commit()
    log_audit(db, current_user, "delete", "master_facility", str(m.id),
              request=request, commit=True)
    return {"success": True}
