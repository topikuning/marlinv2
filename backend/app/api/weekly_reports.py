import os
import tempfile
from decimal import Decimal
from datetime import date, timedelta
from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import io

from app.core.database import get_db
from app.models.models import (
    WeeklyReport, WeeklyProgressItem, WeeklyReportPhoto,
    Contract, BOQItem, Facility, Location, User,
)
from app.schemas.schemas import (
    WeeklyReportCreate, WeeklyReportUpdate, WeeklyReportOut, WeeklyReportDetail,
    ProgressItemInput, ExcelImportResult,
)
from app.api.deps import get_current_user, require_permission, user_can_access_contract
from app.services.audit_service import log_audit
from app.services.progress_service import (
    update_progress_item_calculations, recalculate_report_totals, run_early_warning_check,
    get_previous_cumulatives, compute_facility_progress_summary,
    get_deviation_status, calculate_spi,
)
from app.services.file_service import save_upload, delete_file, ALLOWED_IMAGE_EXT
from app.services.template_service import template_weekly_progress

router = APIRouter(prefix="/reports/weekly", tags=["weekly_reports"])


def _report_to_dict(r: WeeklyReport, detail=False, db: Session = None) -> dict:
    data = {
        "id": str(r.id),
        "contract_id": str(r.contract_id),
        "week_number": r.week_number,
        "period_start": r.period_start.isoformat(),
        "period_end": r.period_end.isoformat(),
        "report_date": r.report_date.isoformat() if r.report_date else None,
        "planned_weekly_pct": float(r.planned_weekly_pct or 0),
        "planned_cumulative_pct": float(r.planned_cumulative_pct or 0),
        "actual_weekly_pct": float(r.actual_weekly_pct or 0),
        "actual_cumulative_pct": float(r.actual_cumulative_pct or 0),
        "deviation_pct": float(r.deviation_pct or 0),
        "deviation_status": r.deviation_status.value if hasattr(r.deviation_status, "value") else r.deviation_status,
        "days_elapsed": r.days_elapsed,
        "days_remaining": r.days_remaining,
        "spi": float(r.spi) if r.spi else None,
        "manpower_count": r.manpower_count,
        "manpower_skilled": r.manpower_skilled,
        "manpower_unskilled": r.manpower_unskilled,
        "rain_days": r.rain_days,
        "obstacles": r.obstacles,
        "solutions": r.solutions,
        "executive_summary": r.executive_summary,
        "submitted_by": r.submitted_by,
        "is_locked": r.is_locked,
        "created_at": r.created_at.isoformat(),
    }
    if detail:
        data["progress_items"] = [
            {
                "id": str(pi.id),
                "boq_item_id": str(pi.boq_item_id),
                "volume_this_week": float(pi.volume_this_week or 0),
                "volume_cumulative": float(pi.volume_cumulative or 0),
                "progress_this_week_pct": float(pi.progress_this_week_pct or 0),
                "progress_cumulative_pct": float(pi.progress_cumulative_pct or 0),
                "weighted_progress_pct": float(pi.weighted_progress_pct or 0),
                "notes": pi.notes,
            }
            for pi in r.progress_items
        ]
        data["photos"] = [
            {
                "id": str(p.id),
                "facility_id": str(p.facility_id) if p.facility_id else None,
                "file_path": p.file_path,
                "thumbnail_path": p.thumbnail_path,
                "caption": p.caption,
                "taken_at": p.taken_at.isoformat() if p.taken_at else None,
                "created_at": p.created_at.isoformat(),
            }
            for p in r.photos
        ]
    return data


# ═══════════════════════════════════════════ LIST / DETAIL ═══════════════════

@router.get("/by-contract/{contract_id}", response_model=dict)
def list_reports(
    contract_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("report.read")),
):
    if not user_can_access_contract(db, current_user, contract_id):
        raise HTTPException(403, "Akses ditolak")
    rows = (
        db.query(WeeklyReport)
        .filter(WeeklyReport.contract_id == contract_id, WeeklyReport.is_deleted == False)
        .order_by(WeeklyReport.week_number.desc())
        .all()
    )
    return {"items": [_report_to_dict(r) for r in rows]}


@router.get("/{report_id}", response_model=dict)
def get_report(
    report_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("report.read")),
):
    r = db.query(WeeklyReport).filter(
        WeeklyReport.id == report_id, WeeklyReport.is_deleted == False
    ).first()
    if not r:
        raise HTTPException(404, "Laporan tidak ditemukan")
    if not user_can_access_contract(db, current_user, str(r.contract_id)):
        raise HTTPException(403, "Akses ditolak")
    return _report_to_dict(r, detail=True, db=db)


# ═══════════════════════════════════════════ CREATE ══════════════════════════

@router.post("/by-contract/{contract_id}", response_model=dict)
def create_report(
    contract_id: str, data: WeeklyReportCreate, request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("report.create")),
):
    if not user_can_access_contract(db, current_user, contract_id):
        raise HTTPException(403, "Akses ditolak")
    c = db.query(Contract).filter(Contract.id == contract_id, Contract.deleted_at.is_(None)).first()
    if not c:
        raise HTTPException(404, "Kontrak tidak ditemukan")

    # Status gate. DRAFT is deliberately allowed (catatan #6) so users can
    # prepare their first report in parallel with contract setup — the
    # report will just display a "contract not yet activated" hint in UI.
    # Completed/terminated contracts are frozen.
    status_value = c.status.value if hasattr(c.status, "value") else str(c.status)
    if status_value in ("completed", "terminated"):
        raise HTTPException(
            400,
            f"Kontrak berstatus '{status_value}' tidak menerima laporan baru.",
        )

    if db.query(WeeklyReport).filter(
        WeeklyReport.contract_id == contract_id,
        WeeklyReport.week_number == data.week_number,
        WeeklyReport.is_deleted == False,
    ).first():
        raise HTTPException(400, f"Laporan minggu ke-{data.week_number} sudah ada. Gunakan edit.")

    days_elapsed = (data.period_end - c.start_date).days + 1 if c.start_date else 0
    days_remaining = (c.end_date - data.period_end).days if c.end_date else 0

    report = WeeklyReport(
        contract_id=contract_id,
        week_number=data.week_number,
        period_start=data.period_start,
        period_end=data.period_end,
        report_date=data.report_date or data.period_end,
        planned_weekly_pct=data.planned_weekly_pct or Decimal("0"),
        planned_cumulative_pct=data.planned_cumulative_pct or Decimal("0"),
        days_elapsed=days_elapsed,
        days_remaining=days_remaining,
        manpower_count=data.manpower_count,
        manpower_skilled=data.manpower_skilled,
        manpower_unskilled=data.manpower_unskilled,
        rain_days=data.rain_days,
        obstacles=data.obstacles,
        solutions=data.solutions,
        executive_summary=data.executive_summary,
        submitted_by=data.submitted_by or current_user.full_name,
        submitted_by_user_id=current_user.id,
        import_source="manual",
    )
    db.add(report)
    db.flush()

    # Pre-lookup Volume Minggu Lalu dari minggu-minggu sebelumnya
    prev_cums = get_previous_cumulatives(
        db, report.contract_id,
        [str(i.boq_item_id) for i in (data.progress_items or [])],
        report.week_number,
    )
    for item_data in data.progress_items:
        boq = db.query(BOQItem).filter(BOQItem.id == item_data.boq_item_id).first()
        if not boq:
            continue
        pi = WeeklyProgressItem(
            weekly_report_id=report.id,
            boq_item_id=item_data.boq_item_id,
            volume_this_week=item_data.volume_this_week,
            notes=item_data.notes,
        )
        try:
            update_progress_item_calculations(
                pi, boq,
                previous_cumulative=prev_cums.get(str(item_data.boq_item_id), 0.0),
            )
        except ValueError as ve:
            db.rollback()
            raise HTTPException(400, {
                "message": str(ve), "code": "weekly_progress_overflow",
            })
        db.add(pi)

    db.flush()
    recalculate_report_totals(db, report)
    db.commit()
    run_early_warning_check(db, contract_id)
    log_audit(db, current_user, "create", "weekly_report", str(report.id),
              changes={"contract_id": contract_id, "week": data.week_number},
              request=request, commit=True)
    return {"id": str(report.id), "success": True}


# ═══════════════════════════════════════════ UPDATE ══════════════════════════

@router.put("/{report_id}", response_model=dict)
def update_report(
    report_id: str, data: WeeklyReportUpdate, request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("report.update")),
):
    r = db.query(WeeklyReport).filter(
        WeeklyReport.id == report_id, WeeklyReport.is_deleted == False,
    ).first()
    if not r:
        raise HTTPException(404, "Laporan tidak ditemukan")
    # Blokir edit saat laporan terkunci, KECUALI permintaan ini memang
    # untuk membuka kunci (is_locked=False). Tanpa pengecualian ini,
    # tombol "Buka Kunci" jadi mustahil ditekan — laporan akan selalu
    # tertolak dengan "Laporan sudah dikunci".
    incoming = data.model_dump(exclude_unset=True)
    unlocking = incoming.get("is_locked") is False
    if r.is_locked and not unlocking:
        raise HTTPException(400, "Laporan sudah dikunci")

    for field in (
        "period_start", "period_end", "planned_weekly_pct", "planned_cumulative_pct",
        "manpower_count", "manpower_skilled", "manpower_unskilled", "rain_days",
        "obstacles", "solutions", "executive_summary", "is_locked",
    ):
        val = getattr(data, field, None)
        if val is not None:
            setattr(r, field, val)

    # Replace progress items if provided
    if data.progress_items is not None:
        db.query(WeeklyProgressItem).filter(
            WeeklyProgressItem.weekly_report_id == r.id
        ).delete()
        db.flush()
        prev_cums = get_previous_cumulatives(
            db, r.contract_id,
            [str(i.boq_item_id) for i in data.progress_items],
            r.week_number,
        )
        for item_data in data.progress_items:
            boq = db.query(BOQItem).filter(BOQItem.id == item_data.boq_item_id).first()
            if not boq:
                continue
            pi = WeeklyProgressItem(
                weekly_report_id=r.id,
                boq_item_id=item_data.boq_item_id,
                volume_this_week=item_data.volume_this_week,
                notes=item_data.notes,
            )
            try:
                update_progress_item_calculations(
                    pi, boq,
                    previous_cumulative=prev_cums.get(str(item_data.boq_item_id), 0.0),
                )
            except ValueError as ve:
                db.rollback()
                raise HTTPException(400, {"message": str(ve), "code": "weekly_progress_overflow"})
            db.add(pi)
    db.flush()
    recalculate_report_totals(db, r)
    db.commit()
    run_early_warning_check(db, str(r.contract_id))
    log_audit(db, current_user, "update", "weekly_report", str(r.id), request=request, commit=True)
    return {"success": True}


# ═══════════════════════════════════════════ UPSERT PROGRESS GRID ════════════

@router.put("/{report_id}/progress-items", response_model=dict)
def upsert_progress_items(
    report_id: str,
    items: List[ProgressItemInput],
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("report.update")),
):
    """
    Partial update: AG Grid posts changed rows only.
    Creates or updates WeeklyProgressItem by (report_id, boq_item_id).
    """
    r = db.query(WeeklyReport).filter(
        WeeklyReport.id == report_id, WeeklyReport.is_deleted == False
    ).first()
    if not r:
        raise HTTPException(404, "Laporan tidak ditemukan")
    if r.is_locked:
        raise HTTPException(400, "Laporan sudah dikunci")

    # Batch-lookup Volume Minggu Lalu untuk semua boq_item_id yang dikirim
    # sekaligus (hindari N+1). volume_cumulative dihitung server: prev+this_week.
    boq_ids = [str(i.boq_item_id) for i in items]
    prev_cums = get_previous_cumulatives(db, r.contract_id, boq_ids, r.week_number)

    touched = 0
    errors = []
    for item_data in items:
        boq = db.query(BOQItem).filter(BOQItem.id == item_data.boq_item_id).first()
        if not boq:
            continue
        pi = db.query(WeeklyProgressItem).filter(
            WeeklyProgressItem.weekly_report_id == r.id,
            WeeklyProgressItem.boq_item_id == item_data.boq_item_id,
        ).first()
        if not pi:
            pi = WeeklyProgressItem(
                weekly_report_id=r.id,
                boq_item_id=item_data.boq_item_id,
            )
            db.add(pi)
        pi.volume_this_week = item_data.volume_this_week
        pi.notes = item_data.notes
        try:
            update_progress_item_calculations(
                pi, boq,
                previous_cumulative=prev_cums.get(str(item_data.boq_item_id), 0.0),
            )
            touched += 1
        except ValueError as ve:
            errors.append({
                "boq_item_id": str(item_data.boq_item_id),
                "message": str(ve),
            })
    if errors:
        db.rollback()
        raise HTTPException(400, {
            "message": "Ada input volume yang melebihi BOQ. Perbaiki sebelum simpan.",
            "code": "weekly_progress_overflow",
            "errors": errors,
        })

    db.flush()
    recalculate_report_totals(db, r)
    db.commit()
    run_early_warning_check(db, str(r.contract_id))
    log_audit(db, current_user, "upsert_grid", "weekly_progress_items",
              str(r.id), changes={"touched": touched}, request=request, commit=True)
    return {"success": True, "touched": touched,
            "actual_cumulative_pct": float(r.actual_cumulative_pct or 0),
            "deviation_pct": float(r.deviation_pct or 0),
            "spi": float(r.spi or 0)}


# ═══════════════════════════════════════════ DELETE ══════════════════════════

@router.delete("/{report_id}", response_model=dict)
def delete_report(
    report_id: str, request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("report.delete")),
):
    r = db.query(WeeklyReport).filter(WeeklyReport.id == report_id).first()
    if not r:
        raise HTTPException(404, "Laporan tidak ditemukan")
    r.is_deleted = True
    db.commit()
    log_audit(db, current_user, "delete", "weekly_report", str(r.id), request=request, commit=True)
    return {"success": True}


# ═══════════════════════════════════════════ PHOTOS ══════════════════════════

@router.post("/{report_id}/photos", response_model=dict)
async def upload_photo(
    report_id: str,
    file: UploadFile = File(...),
    caption: Optional[str] = Form(None),
    facility_id: Optional[str] = Form(None),
    request: Request = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("report.update")),
):
    r = db.query(WeeklyReport).filter(
        WeeklyReport.id == report_id, WeeklyReport.is_deleted == False
    ).first()
    if not r:
        raise HTTPException(404, "Laporan tidak ditemukan")
    rel, thumb = save_upload(file, "weekly", ALLOWED_IMAGE_EXT)
    p = WeeklyReportPhoto(
        weekly_report_id=r.id,
        facility_id=facility_id if facility_id else None,
        file_path=rel,
        thumbnail_path=thumb,
        caption=caption,
        uploaded_by=current_user.id,
    )
    db.add(p)
    db.commit()
    db.refresh(p)
    return {
        "id": str(p.id),
        "file_path": rel,
        "thumbnail_path": thumb,
        "caption": caption,
    }


@router.delete("/{report_id}/photos/{photo_id}", response_model=dict)
def delete_photo(
    report_id: str, photo_id: str, request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("report.update")),
):
    p = db.query(WeeklyReportPhoto).filter(
        WeeklyReportPhoto.id == photo_id, WeeklyReportPhoto.weekly_report_id == report_id,
    ).first()
    if not p:
        raise HTTPException(404, "Foto tidak ditemukan")
    delete_file(p.file_path)
    delete_file(p.thumbnail_path)
    db.delete(p)
    db.commit()
    return {"success": True}


# ═══════════════════════════════════════════ TEMPLATE ════════════════════════

@router.get("/template/{contract_id}")
def download_progress_template(
    contract_id: str,
    db: Session = Depends(get_db),
    _=Depends(require_permission("report.read")),
):
    items = (
        db.query(BOQItem, Facility, Location)
        .join(Facility, Facility.id == BOQItem.facility_id)
        .join(Location, Location.id == Facility.location_id)
        .filter(Location.contract_id == contract_id,
                BOQItem.is_active == True, BOQItem.is_leaf == True)
        .order_by(Location.location_code, Facility.display_order, BOQItem.display_order)
        .all()
    )
    rows = [
        {
            "id": str(b.id),
            "full_code": b.full_code,
            "original_code": b.original_code,
            "description": f"[{l.location_code} / {f.facility_code}] {b.description}",
            "unit": b.unit,
            "volume": float(b.volume or 0),
        }
        for b, f, l in items
    ]
    data = template_weekly_progress(rows)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=template_progress_{contract_id[:8]}.xlsx"},
    )


# ─────────────────────────────────────────────────────────────────────────────
# Excel export / import untuk progress per minggu (best-practice bulk edit
# di Excel, upload balik → server upsert by boq_item_id).
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/{report_id}/export-excel")
def export_report_excel(
    report_id: str,
    db: Session = Depends(get_db),
    _=Depends(require_permission("report.read")),
):
    r = db.query(WeeklyReport).filter(
        WeeklyReport.id == report_id, WeeklyReport.is_deleted == False,
    ).first()
    if not r:
        raise HTTPException(404, "Laporan tidak ditemukan")

    # Ambil daftar BOQ item lengkap dengan lokasi/fasilitas, lalu left-join
    # progress yang sudah diinput untuk laporan ini.
    boq_rows = (
        db.query(BOQItem, Facility, Location)
        .join(Facility, Facility.id == BOQItem.facility_id)
        .join(Location, Location.id == Facility.location_id)
        .filter(
            Location.contract_id == r.contract_id,
            BOQItem.is_active == True,  # noqa: E712
            BOQItem.is_leaf == True,    # noqa: E712
        )
        .order_by(Location.location_code, Facility.display_order, BOQItem.display_order)
        .all()
    )
    progress_map = {
        p.boq_item_id: p
        for p in db.query(WeeklyProgressItem).filter(
            WeeklyProgressItem.weekly_report_id == r.id,
        ).all()
    }

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Progress"

    headers = [
        "boq_item_id", "Kode Lokasi", "Kode Fasilitas", "Nama Fasilitas",
        "Kode BOQ", "Uraian", "Satuan", "Vol BOQ", "Bobot %",
        "Vol Minggu Ini", "Vol Kumulatif", "Progress Kum %", "Catatan",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        c.alignment = Alignment(horizontal="center")

    for row_i, (b, f, l) in enumerate(boq_rows, start=2):
        p = progress_map.get(b.id)
        ws.cell(row=row_i, column=1, value=str(b.id))
        ws.cell(row=row_i, column=2, value=l.location_code)
        ws.cell(row=row_i, column=3, value=f.facility_code)
        ws.cell(row=row_i, column=4, value=f.facility_name)
        ws.cell(row=row_i, column=5, value=b.full_code or b.original_code or "")
        ws.cell(row=row_i, column=6, value=b.description or "")
        ws.cell(row=row_i, column=7, value=b.unit or "")
        ws.cell(row=row_i, column=8, value=float(b.volume or 0))
        ws.cell(row=row_i, column=9, value=float(b.weight_pct or 0) * 100)
        ws.cell(row=row_i, column=10, value=float(p.volume_this_week) if p else 0)
        ws.cell(row=row_i, column=11, value=float(p.volume_cumulative) if p else 0)
        ws.cell(row=row_i, column=12, value=float(p.progress_cumulative_pct) * 100 if p else 0)
        ws.cell(row=row_i, column=13, value=(p.notes if p else "") or "")

    # Auto-width approximations
    widths = [36, 12, 14, 28, 14, 48, 10, 10, 10, 14, 14, 14, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w
    ws.freeze_panes = "E2"

    # Petunjuk di sheet kedua
    ws2 = wb.create_sheet("Petunjuk")
    tips = [
        "PETUNJUK PENGISIAN",
        "",
        "1. JANGAN hapus atau ubah kolom 'boq_item_id' (kolom A). Server memakai kolom ini untuk mencocokkan data saat upload.",
        "2. Isi kolom 'Vol Minggu Ini' (kolom J) dengan volume yang dikerjakan minggu ini saja.",
        "3. Isi kolom 'Vol Kumulatif' (kolom K) dengan total akumulasi sampai minggu ini.",
        "4. Kolom 'Progress Kum %' (kolom L) akan dihitung ulang otomatis saat upload — tidak perlu diisi manual.",
        "5. Kolom 'Catatan' (kolom M) opsional, terpakai bila ada hal yang perlu diinfokan.",
        "6. Simpan file sebagai .xlsx, lalu upload kembali via tombol 'Import Excel' di halaman laporan.",
        "",
        "Baris yang tidak diubah boleh dibiarkan apa adanya — server mengupsert berdasar boq_item_id.",
    ]
    for i, t in enumerate(tips, start=1):
        c = ws2.cell(row=i, column=1, value=t)
        if i == 1:
            c.font = Font(bold=True, size=14)
    ws2.column_dimensions["A"].width = 110

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"progress_W{r.week_number}_{r.contract_id.hex[:8]}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/{report_id}/import-excel", response_model=dict)
async def import_report_excel(
    report_id: str,
    file: UploadFile = File(...),
    request: Request = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("report.update")),
):
    r = db.query(WeeklyReport).filter(
        WeeklyReport.id == report_id, WeeklyReport.is_deleted == False,
    ).first()
    if not r:
        raise HTTPException(404, "Laporan tidak ditemukan")
    if r.is_locked:
        raise HTTPException(400, "Laporan sudah dikunci — buka kunci sebelum import.")

    import pandas as pd
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    result = {"updated": 0, "created": 0, "skipped": 0, "errors": []}
    try:
        df = pd.read_excel(tmp_path, sheet_name="Progress", dtype=object)
        if "boq_item_id" not in [str(c).strip() for c in df.columns]:
            result["errors"].append("Kolom 'boq_item_id' tidak ditemukan. Pastikan pakai file hasil Export Excel.")
            return result

        # Batch lookup Volume Minggu Lalu untuk semua boq_item_id di file
        file_boq_ids = [
            str(row.get("boq_item_id") or "").strip()
            for _, row in df.iterrows()
            if str(row.get("boq_item_id") or "").strip()
        ]
        prev_cums = get_previous_cumulatives(db, r.contract_id, file_boq_ids, r.week_number)

        for idx, row in df.iterrows():
            bid = str(row.get("boq_item_id") or "").strip()
            if not bid or bid.lower() == "nan":
                result["skipped"] += 1
                continue
            boq = db.query(BOQItem).filter(BOQItem.id == bid).first()
            if not boq:
                result["skipped"] += 1
                continue
            def _f(v):
                try: return float(v)
                except (TypeError, ValueError): return 0.0

            vol_week = _f(row.get("Vol Minggu Ini"))
            vol_cum = _f(row.get("Vol Kumulatif"))
            notes = str(row.get("Catatan") or "") or None

            pi = db.query(WeeklyProgressItem).filter(
                WeeklyProgressItem.weekly_report_id == r.id,
                WeeklyProgressItem.boq_item_id == bid,
            ).first()
            if not pi:
                pi = WeeklyProgressItem(
                    weekly_report_id=r.id,
                    boq_item_id=bid,
                )
                db.add(pi)
                result["created"] += 1
            else:
                result["updated"] += 1
            pi.volume_this_week = Decimal(str(vol_week))
            pi.notes = notes
            # Abaikan Vol Kumulatif di Excel — server yang compute ulang
            # (prev + this_week) untuk menjaga konsistensi histori.
            try:
                update_progress_item_calculations(
                    pi, boq,
                    previous_cumulative=prev_cums.get(str(bid), 0.0),
                )
            except ValueError as ve:
                result["errors"].append(f"Baris {idx+2} ({boq.description or bid}): {ve}")
                continue

        db.flush()
        recalculate_report_totals(db, r)
        db.commit()
        log_audit(
            db, current_user, "import_excel", "weekly_report", str(r.id),
            changes=result, request=request, commit=True,
        )
    except Exception as e:
        result["errors"].append(str(e))
    finally:
        try: os.remove(tmp_path)
        except OSError: pass
    result["success"] = not result["errors"]
    return result
