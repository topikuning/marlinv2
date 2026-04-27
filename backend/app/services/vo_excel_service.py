"""
Excel snapshot + import untuk VO bulk edit.

Alur (Konvensi B):
  - Export: BOQ aktif per facility + ringkasan VO APPROVED-unbundled lain
  - User edit kolom 'vol_baru' di Excel
  - Import: hitung delta = vol_baru - vol_efektif_snapshot per item, infer
    action otomatis, return list VOItemInput yg client gabung ke form

Konvensi: vol_baru = volume final yg user inginkan dengan asumsi semua VO
APPROVED-pending (vol_pending_vo_lain) ikut di-bundle. delta yg disimpan
relatif ke vol_efektif, bukan vol_awal.

Mode export:
  - "flat"          → 1 sheet BOQ_Snapshot berisi semua fasilitas (default, backward-compat)
  - "per_facility"  → 1 sheet per fasilitas (FAC_<kode>) + sheet REKAP + sheet Petunjuk

Mode import: auto-detect dari sheet names:
  - Ada sheet FAC_*  → multi-sheet mode (parse setiap FAC_* sheet)
  - Tidak ada        → flat mode (baca sheet pertama)
"""
from __future__ import annotations
import io
import re
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict, List, Optional, Set, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session

from app.models.models import (
    Contract, BOQItem, BOQRevision, RevisionStatus, Facility, Location,
    VariationOrder, VariationOrderItem, VOStatus, VOItemAction,
)


HEADERS = [
    "boq_item_id",  # PRIMARY MATCHING KEY — jangan diubah/dihapus
    "facility_code", "facility_name",
    "code", "parent_code",
    "description", "unit",
    "vol_awal", "vol_pending_vo_lain", "nilai_pending",
    "vol_efektif", "vol_baru",
    "unit_price", "catatan_vo_lain",
]


# Whitelist keyword di awal kolom catatan_vo_lain untuk membypass validasi
# unit_price=0 saat ADD. Format: "<KEYWORD>" atau "<KEYWORD>: <alasan>"
# (case-insensitive, whitespace boleh setelah keyword).
ZERO_PRICE_BYPASS_KEYWORDS = ("PARENT", "INFO", "OWNER", "TITIPAN")


def _check_zero_price_bypass(notes: str) -> bool:
    """True jika `notes` diawali salah satu keyword whitelist (case-insensitive)."""
    if not notes:
        return False
    text = notes.strip().upper()
    for kw in ZERO_PRICE_BYPASS_KEYWORDS:
        if text == kw or text.startswith(f"{kw}:") or text.startswith(f"{kw} "):
            return True
    return False


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _active_revision(db: Session, contract_id) -> Optional[BOQRevision]:
    return (
        db.query(BOQRevision)
        .filter(
            BOQRevision.contract_id == contract_id,
            BOQRevision.is_active == True,  # noqa: E712
            BOQRevision.status == RevisionStatus.APPROVED,
        )
        .first()
    )


def _pending_for_item(
    db: Session, boq_item_id, exclude_vo_id=None
) -> Tuple[Decimal, Decimal, List[str]]:
    """
    Ringkasan VO APPROVED-unbundled lain yang menyentuh item ini.

    Return: (sum_volume_delta, sum_cost_impact, ['VO-001: +5 m³', ...]).
    Hanya hitung action yang ubah volume (INCREASE/DECREASE/REMOVE).
    """
    q = (
        db.query(VariationOrderItem, VariationOrder)
        .join(VariationOrder, VariationOrder.id == VariationOrderItem.variation_order_id)
        .filter(
            VariationOrderItem.boq_item_id == boq_item_id,
            VariationOrder.status == VOStatus.APPROVED,
            VariationOrder.bundled_addendum_id.is_(None),
        )
    )
    if exclude_vo_id:
        q = q.filter(VariationOrder.id != exclude_vo_id)

    total_vol = Decimal("0")
    total_cost = Decimal("0")
    notes: List[str] = []
    for vi, vo in q.all():
        if vi.action not in (VOItemAction.INCREASE, VOItemAction.DECREASE, VOItemAction.REMOVE):
            continue
        d = Decimal(vi.volume_delta or 0)
        if vi.action == VOItemAction.REMOVE:
            d = -Decimal(vi.volume_delta or 0)  # delta sudah negatif kalau remove
            # actually for REMOVE we typically store delta = -original. Re-fetch
            # safe via cost_impact sign. Just use vi.volume_delta as-is.
            d = Decimal(vi.volume_delta or 0)
        total_vol += d
        total_cost += Decimal(vi.cost_impact or 0)
        sign = "+" if d >= 0 else ""
        notes.append(f"{vo.vo_number}: {sign}{d} {vi.unit or ''}")
    return total_vol, total_cost, notes


def _pending_remove_facility(
    db: Session, facility_id, exclude_vo_id=None
) -> List[VariationOrder]:
    """List VO APPROVED-unbundled yg menandai REMOVE_FACILITY untuk facility ini."""
    q = (
        db.query(VariationOrder)
        .join(VariationOrderItem, VariationOrder.id == VariationOrderItem.variation_order_id)
        .filter(
            VariationOrderItem.facility_id == facility_id,
            VariationOrderItem.action == VOItemAction.REMOVE_FACILITY,
            VariationOrder.status == VOStatus.APPROVED,
            VariationOrder.bundled_addendum_id.is_(None),
        )
        .distinct()
    )
    if exclude_vo_id:
        q = q.filter(VariationOrder.id != exclude_vo_id)
    return q.all()


def _excel_sheet_name(facility_code: str) -> str:
    """Buat nama sheet Excel yang valid untuk fasilitas (max 31 char, tanpa karakter terlarang)."""
    name = f"FAC_{facility_code}"
    # Ganti karakter terlarang Excel: / \ * ? : [ ]
    name = re.sub(r"[/\\*?:\[\]]", "_", name)
    return name[:31]


def _setup_ws_header(ws) -> None:
    """Tulis baris header HEADERS ke worksheet dengan styling."""
    ws.append(HEADERS)
    fill = PatternFill("solid", fgColor="0F172A")
    font_h = Font(color="FFFFFF", bold=True, size=10)
    for c in ws[1]:
        c.fill = fill
        c.font = font_h
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _apply_ws_styling(ws) -> None:
    """Sembunyikan kolom UUID (A) dan auto-width kolom lain."""
    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["A"].width = 0
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col_letter == "A":
            continue
        max_len = 8
        for cell in col:
            if cell.value is None:
                continue
            v = str(cell.value)
            if len(v) > max_len:
                max_len = min(len(v), 60)
        ws.column_dimensions[col_letter].width = max_len + 2


def _write_facility_rows(
    ws,
    fac: Facility,
    items: List[BOQItem],
    code_by_id: Dict[str, str],
    edit_vo_items_by_boq: Dict[str, VariationOrderItem],
    exclude_vo_id,
    db: Session,
) -> None:
    """Tulis baris BOQ satu fasilitas ke worksheet (tanpa header)."""
    fac_remove_vos = _pending_remove_facility(db, fac.id, exclude_vo_id=exclude_vo_id)
    fac_remove_note = (
        "; ".join([f"{vo.vo_number}: HILANGKAN FASILITAS" for vo in fac_remove_vos])
        if fac_remove_vos else ""
    )

    for it in items:
        parent_code_str = code_by_id.get(str(it.parent_id), "") if it.parent_id else ""
        if not it.is_leaf:
            ws.append([
                str(it.id),
                fac.facility_code, fac.facility_name,
                it.original_code or "", parent_code_str,
                it.description, it.unit or "",
                "", "", "", "", "",
                "", fac_remove_note or "(group/parent — vol_baru kosong)",
            ])
            continue

        vol_awal = Decimal(it.volume or 0)
        unit_price = Decimal(it.unit_price or 0)
        total_price = Decimal(it.total_price or 0)
        sum_vol, sum_cost, notes_item = _pending_for_item(db, it.id, exclude_vo_id=exclude_vo_id)

        if fac_remove_vos:
            pending_vol = -vol_awal
            pending_cost = -total_price
            combined_notes = "; ".join([fac_remove_note] + notes_item)
        else:
            pending_vol = sum_vol
            pending_cost = sum_cost
            combined_notes = "; ".join(notes_item) if notes_item else ""

        vol_efektif = vol_awal + pending_vol

        existing = edit_vo_items_by_boq.get(str(it.id))
        if existing:
            d_this = Decimal(existing.volume_delta or 0)
            vol_baru = vol_efektif + d_this
        else:
            vol_baru = vol_efektif

        ws.append([
            str(it.id),
            fac.facility_code, fac.facility_name,
            it.original_code or "",
            parent_code_str,
            it.description, it.unit or "",
            float(_q2(vol_awal)),
            float(_q2(pending_vol)),
            float(_q2(pending_cost)),
            float(_q2(vol_efektif)),
            float(_q2(vol_baru)),
            float(_q2(unit_price)),
            combined_notes,
        ])


def _add_petunjuk_sheet(wb: Workbook, mode: str = "flat") -> None:
    ws = wb.create_sheet("Petunjuk")
    ws["A1"] = "PETUNJUK PENGISIAN BOQ SNAPSHOT (untuk VO Bulk Edit)"
    ws["A1"].font = Font(bold=True, size=14)

    if mode == "per_facility":
        mode_info = [
            "",
            f"Format file ini: PER FASILITAS — setiap sheet 'FAC_<kode>' berisi item satu fasilitas.",
            "Sheet REKAP: daftar fasilitas (hanya referensi, tidak dibaca saat import).",
            "Saat import, sistem otomatis deteksi sheet FAC_* dan parse semua sekaligus.",
            "Anda bisa menghapus sheet FAC_* yang tidak perlu diedit sebelum upload.",
        ]
    else:
        mode_info = [
            "",
            "Format file ini: FLAT — semua fasilitas ada di sheet 'BOQ_Snapshot'.",
        ]

    instr = mode_info + [
        "",
        "KONSEP:",
        "Anda terima snapshot BOQ aktif + info VO lain yang sudah APPROVED",
        "tapi belum di-bundle. Anda edit kolom 'vol_baru' saja, sistem hitung delta",
        "otomatis saat upload kembali.",
        "",
        "KOLOM:",
        "- boq_item_id (HIDDEN, kolom A): kunci matching ke item BOQ existing.",
        "  JANGAN diubah/dihapus/edit. Kalau hilang, sistem tidak bisa match",
        "  dan akan menganggap row sebagai item baru (ADD).",
        "- facility_code, facility_name, code, parent_code, description, unit: identifier (read-only)",
        "- vol_awal: volume di revisi aktif kontrak (read-only)",
        "- vol_pending_vo_lain: total Δ volume dari VO APPROVED lain yang menyentuh item ini",
        "- nilai_pending: total Δ Rp dari VO lain itu",
        "- vol_efektif: vol_awal + vol_pending — proyeksi kalau semua VO pending lolos",
        "- vol_baru: ★ KOLOM YG ANDA EDIT ★ — volume final yang anda inginkan",
        "  Default = vol_efektif (artinya tidak ada perubahan baru dari VO ini)",
        "- unit_price: harga satuan (read-only — kontrak tidak boleh diubah)",
        "- catatan_vo_lain: daftar VO lain yang sudah ubah item ini",
        "",
        "AKSI OTOMATIS:",
        "- vol_baru > vol_efektif → INCREASE (Δ = vol_baru - vol_efektif)",
        "- vol_baru < vol_efektif (tapi > 0) → DECREASE",
        "- vol_baru = 0 → REMOVE (item dihapus)",
        "- Tambah baris baru (kosongkan boq_item_id) → ADD",
        "  - Untuk ADD, isi: facility_code, parent_code (opsional), description, unit, vol_baru, unit_price",
        "  - vol_efektif & vol_pending biarkan kosong",
        "  - parent_code boleh merujuk ke 'code' baris ADD lain di sheet ini",
        "    (chain ADD → ADD didukung; sistem otomatis urut hirarki saat apply)",
        "",
        "ADD DENGAN unit_price = 0 (proteksi typo, default ditolak):",
        "Untuk item BOQ yang memang berharga 0 secara sah, isi kolom 'catatan_vo_lain'",
        "diawali salah satu keyword berikut + opsional alasan:",
        "  - PARENT   → item agregator/header hirarki (group, parent dari child ADD)",
        "  - INFO     → item informatif/non-cost (referensi koordinasi)",
        "  - OWNER    → material/jasa disediakan KKP (owner-supplied)",
        "  - TITIPAN  → biaya ditanggung pihak lain",
        "Contoh: 'PARENT: Pondasi tambahan blok B' / 'OWNER: Pompa air supplied KKP'",
        "Tanpa keyword whitelist, baris dengan unit_price=0 akan ditolak.",
        "",
        "EDGE CASES:",
        "- Vol_baru < 0: tidak diizinkan, akan ditolak saat upload.",
        "- Unit_price < 0: tidak diizinkan.",
        "- Group row (parent, is_leaf=false): biarkan vol_baru kosong; tidak ditolak.",
        "- Code tidak ditemukan saat upload: dianggap baris ADD (item baru).",
        "- VO referensi di-reject oleh PPK: VO Anda tetap valid tapi hasilnya tidak match.",
        "  Sistem kasih warning di addendum modal saat bundle.",
    ]
    for i, line in enumerate(instr, start=2):
        ws[f"A{i}"] = line
    ws.column_dimensions["A"].width = 100


# ─── Export ───────────────────────────────────────────────────────────────────

def export_snapshot(
    db: Session,
    contract_id: str,
    facility_ids: Optional[List[str]] = None,
    exclude_vo_id: Optional[str] = None,
    mode: str = "flat",
) -> bytes:
    """
    Generate Excel snapshot.

    facility_ids=None → semua facility kontrak.
    exclude_vo_id     → VO yang sedang di-edit, item-nya di-skip dari pending sum
                        dan vol_baru di-prefill dari delta VO ini.
    mode              → "flat" (1 sheet, default) | "per_facility" (sheet per fasilitas)
    """
    rev = _active_revision(db, contract_id)
    if not rev:
        rev = (
            db.query(BOQRevision)
            .filter(BOQRevision.contract_id == contract_id)
            .order_by(BOQRevision.cco_number.desc())
            .first()
        )
    if not rev:
        raise ValueError("Kontrak belum punya revisi BOQ.")

    # Resolve facilities
    fac_q = (
        db.query(Facility, Location)
        .join(Location, Facility.location_id == Location.id)
        .filter(Location.contract_id == contract_id)
    )
    if facility_ids:
        fac_q = fac_q.filter(Facility.id.in_(facility_ids))
    facilities = fac_q.order_by(Location.location_code, Facility.facility_code).all()

    # VO yang sedang di-edit — ambil itemnya untuk pre-fill vol_baru
    edit_vo_items_by_boq: Dict[str, VariationOrderItem] = {}
    if exclude_vo_id:
        for vi in db.query(VariationOrderItem).filter(
            VariationOrderItem.variation_order_id == exclude_vo_id
        ).all():
            if vi.boq_item_id:
                edit_vo_items_by_boq[str(vi.boq_item_id)] = vi

    wb = Workbook()

    if mode == "per_facility":
        # ── REKAP sheet ──────────────────────────────────────────────────────
        ws_rekap = wb.active
        ws_rekap.title = "REKAP"
        rekap_headers = ["No.", "Lokasi", "Kode Fasilitas", "Nama Fasilitas", "Jumlah Item Leaf", "Sheet"]
        ws_rekap.append(rekap_headers)
        fill_hdr = PatternFill("solid", fgColor="0F172A")
        font_hdr = Font(color="FFFFFF", bold=True, size=10)
        for c in ws_rekap[1]:
            c.fill = fill_hdr
            c.font = font_hdr
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_rekap.freeze_panes = "A2"

        for i, (fac, loc) in enumerate(facilities, start=1):
            items = (
                db.query(BOQItem)
                .filter(
                    BOQItem.boq_revision_id == rev.id,
                    BOQItem.facility_id == fac.id,
                    BOQItem.is_active == True,  # noqa: E712
                )
                .order_by(BOQItem.display_order, BOQItem.full_code)
                .all()
            )
            leaf_count = sum(1 for it in items if it.is_leaf)
            sheet_name = _excel_sheet_name(fac.facility_code)
            ws_rekap.append([i, loc.location_code, fac.facility_code, fac.facility_name, leaf_count, sheet_name])

            # ── Sheet per fasilitas ──────────────────────────────────────────
            ws_fac = wb.create_sheet(sheet_name)
            _setup_ws_header(ws_fac)
            code_by_id: Dict[str, str] = {str(it.id): (it.original_code or "") for it in items}
            _write_facility_rows(ws_fac, fac, items, code_by_id, edit_vo_items_by_boq, exclude_vo_id, db)
            _apply_ws_styling(ws_fac)

        # Auto-width REKAP
        for col in ws_rekap.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws_rekap.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    else:
        # ── Flat mode: single BOQ_Snapshot sheet (backward-compat) ──────────
        ws = wb.active
        ws.title = "BOQ_Snapshot"
        _setup_ws_header(ws)

        for fac, loc in facilities:
            items = (
                db.query(BOQItem)
                .filter(
                    BOQItem.boq_revision_id == rev.id,
                    BOQItem.facility_id == fac.id,
                    BOQItem.is_active == True,  # noqa: E712
                )
                .order_by(BOQItem.display_order, BOQItem.full_code)
                .all()
            )
            code_by_id: Dict[str, str] = {str(it.id): (it.original_code or "") for it in items}
            _write_facility_rows(ws, fac, items, code_by_id, edit_vo_items_by_boq, exclude_vo_id, db)

        _apply_ws_styling(ws)

    _add_petunjuk_sheet(wb, mode=mode)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Parse ────────────────────────────────────────────────────────────────────

def _parse_df_rows(
    df: "pd.DataFrame",
    boq_by_uuid: Dict[str, BOQItem],
    boq_by_code: Dict[Tuple[str, str], BOQItem],
    fac_index: Dict[str, Facility],
    exclude_vo_id,
    db: Session,
) -> Tuple[List[Dict], List[str], List[str], Set[str]]:
    """
    Parse satu DataFrame (satu sheet) snapshot VO.
    Return: (items_out, warnings, errors, facility_codes_in_sheet)
    """
    cols = [str(c).strip().lower() for c in df.columns]
    required = ["facility_code", "vol_baru"]
    missing = [c for c in required if c not in cols]
    if missing:
        return [], [], [f"Kolom wajib hilang: {', '.join(missing)}"], set()

    has_uuid_col = "boq_item_id" in cols
    if not has_uuid_col:
        return [], [], [
            "Kolom 'boq_item_id' tidak ada. File ini bukan dari Export Excel "
            "yang baru — sistem tidak bisa match ke item BOQ existing. "
            "Download Snapshot baru dan edit ulang."
        ], set()

    items_out: List[Dict] = []
    warnings: List[str] = []
    errors: List[str] = []
    fac_codes_in_sheet: Set[str] = set()

    for idx, row in df.iterrows():
        rec = {cols[i]: row.iloc[i] for i in range(len(cols))}
        fac_code = _safe_str(rec.get("facility_code"))
        code = _safe_str(rec.get("code"))
        uuid_str = _safe_str(rec.get("boq_item_id"))
        if not fac_code:
            continue
        fac_codes_in_sheet.add(fac_code)

        fac = fac_index.get(fac_code)
        if not fac:
            errors.append(f"Baris {idx+2}: facility_code '{fac_code}' tidak ada di kontrak.")
            continue

        vol_baru_raw = rec.get("vol_baru")
        if vol_baru_raw is None or (isinstance(vol_baru_raw, str) and not vol_baru_raw.strip()):
            # Skip row tanpa vol_baru (mis. group row)
            continue
        try:
            vol_baru = float(vol_baru_raw)
        except (TypeError, ValueError):
            errors.append(f"Baris {idx+2}: vol_baru '{vol_baru_raw}' bukan angka.")
            continue
        if vol_baru < 0:
            errors.append(f"Baris {idx+2} ({fac_code}/{code}): vol_baru < 0 ditolak.")
            continue

        unit = _safe_str(rec.get("unit"))
        unit_price = _safe_float(rec.get("unit_price"))
        description = _safe_str(rec.get("description"))
        parent_code = _safe_str(rec.get("parent_code"))

        # MATCHING STRATEGY:
        #   1. boq_item_id (UUID) — primary, deterministic
        #   2. Fallback: (facility_code, code) untuk row yang user tambahkan
        #      manual atau row hasil ADD lama
        #   3. Tidak ada match → ADD baru
        boq_item = None
        if uuid_str:
            boq_item = boq_by_uuid.get(uuid_str)
        if not boq_item and code:
            boq_item = boq_by_code.get((fac_code, code))

        if not boq_item:
            # ADD — item baru (UUID kosong DAN code tidak match item existing)
            if vol_baru <= 0:
                continue  # skip empty add row
            if not description:
                errors.append(f"Baris {idx+2}: ADD perlu description.")
                continue
            notes_str = _safe_str(rec.get("catatan_vo_lain"))
            if unit_price < 0:
                errors.append(f"Baris {idx+2}: ADD unit_price tidak boleh negatif.")
                continue
            if unit_price == 0 and not _check_zero_price_bypass(notes_str):
                errors.append(
                    f"Baris {idx+2}: ADD dengan unit_price=0 wajib isi 'catatan_vo_lain' "
                    f"diawali keyword PARENT / INFO / OWNER / TITIPAN "
                    f"(mis. 'PARENT: header pondasi blok B'). "
                    f"Kalau ini bukan disengaja, isi unit_price > 0."
                )
                continue
            # Parent lookup: coba match ke item existing di DB
            # Kalau tidak ketemu → simpan parent_code string untuk di-resolve
            # saat apply (bisa jadi item ADD lain di sheet yang sama)
            parent_boq_id = None
            parent_code_to_store = None
            if parent_code:
                parent_item = boq_by_code.get((fac_code, parent_code))
                if parent_item:
                    parent_boq_id = str(parent_item.id)
                else:
                    # Parent mungkin item ADD baru lain di sheet ini — simpan
                    # sebagai string, tidak lagi jadi warning
                    parent_code_to_store = parent_code
            items_out.append({
                "action": "add",
                "facility_id": str(fac.id),
                "boq_item_id": None,
                "parent_boq_item_id": parent_boq_id,
                "parent_code": parent_code_to_store,
                "new_item_code": code or None,
                "description": description,
                "unit": unit,
                "volume_delta": float(_q2(vol_baru)),
                "unit_price": float(_q2(unit_price)),
                "notes": notes_str or None,
            })
            continue

        # Aturan sistem: volume & unit_price quantized ke 2 dp. Vol_baru di
        # Excel = volume aktif di DB persis (no float noise).
        sum_vol_now, _, _ = _pending_for_item(db, boq_item.id, exclude_vo_id=exclude_vo_id)
        vol_baru_d = _q2(vol_baru)
        vol_awal_d = _q2(boq_item.volume or 0)
        vol_efektif_now_d = _q2(vol_awal_d + sum_vol_now)

        vol_efektif_excel_d = _safe_decimal(rec.get("vol_efektif"))
        if vol_efektif_excel_d is not None:
            vol_efektif_excel_d = _q2(vol_efektif_excel_d)
        # Note: pakai `is not None`, BUKAN truthy — vol_efektif=0 itu nilai sah
        # untuk item baru atau item yang sudah di-cancel oleh VO lain.
        vol_efektif_used_d = (
            vol_efektif_excel_d if vol_efektif_excel_d is not None else vol_efektif_now_d
        )
        if vol_efektif_excel_d is not None and abs(vol_efektif_excel_d - vol_efektif_now_d) > _TWOPLACES:
            warnings.append(
                f"Baris {idx+2} ({fac_code}/{code}): vol_efektif berubah sejak export "
                f"({vol_efektif_excel_d} → {vol_efektif_now_d}). Disarankan re-export."
            )

        delta_d = _q2(vol_baru_d - vol_efektif_used_d)
        if delta_d == Decimal("0.00"):
            continue  # no change
        delta = float(delta_d)
        vol_awal = float(vol_awal_d)
        unit_price_d = _q2(boq_item.unit_price or 0)

        if vol_baru_d == Decimal("0.00") and vol_awal_d > 0:
            # REMOVE — delta = -vol_awal (kembalikan dari rev aktif)
            items_out.append({
                "action": "remove",
                "facility_id": str(boq_item.facility_id),
                "boq_item_id": str(boq_item.id),
                "description": boq_item.description,
                "unit": boq_item.unit,
                "volume_delta": float(-vol_awal_d),
                "unit_price": float(unit_price_d),
            })
        elif delta_d > 0:
            items_out.append({
                "action": "increase",
                "facility_id": str(boq_item.facility_id),
                "boq_item_id": str(boq_item.id),
                "description": boq_item.description,
                "unit": boq_item.unit,
                "volume_delta": delta,
                "unit_price": float(unit_price_d),
            })
        elif delta_d < 0:
            items_out.append({
                "action": "decrease",
                "facility_id": str(boq_item.facility_id),
                "boq_item_id": str(boq_item.id),
                "description": boq_item.description,
                "unit": boq_item.unit,
                "volume_delta": delta,
                "unit_price": float(unit_price_d),
            })

    return items_out, warnings, errors, fac_codes_in_sheet


def parse_snapshot(
    db: Session,
    contract_id: str,
    file_bytes: bytes,
    exclude_vo_id: Optional[str] = None,
) -> Dict:
    """
    Parse Excel hasil export. Return list VOItemInput dicts + warnings.

    Auto-detect format:
    - Sheet FAC_* ada  → multi-sheet per fasilitas (parse semua FAC_* sheet)
    - Tidak ada        → flat (baca sheet pertama)

    Hasil dipakai client untuk replace items di form VO. Tidak menyentuh DB.
    """
    rev = _active_revision(db, contract_id)
    if not rev:
        rev = (
            db.query(BOQRevision)
            .filter(BOQRevision.contract_id == contract_id)
            .order_by(BOQRevision.cco_number.desc())
            .first()
        )
    if not rev:
        raise ValueError("Kontrak belum punya revisi BOQ.")

    # Build indexes untuk matching
    fac_index: Dict[str, Facility] = {}
    fac_by_id: Dict[str, Facility] = {}
    for fac, loc in (
        db.query(Facility, Location)
        .join(Location, Facility.location_id == Location.id)
        .filter(Location.contract_id == contract_id)
        .all()
    ):
        fac_index[fac.facility_code] = fac
        fac_by_id[str(fac.id)] = fac

    boq_by_uuid: Dict[str, BOQItem] = {}
    boq_by_code: Dict[Tuple[str, str], BOQItem] = {}
    for it in (
        db.query(BOQItem)
        .filter(
            BOQItem.boq_revision_id == rev.id,
            BOQItem.is_active == True,  # noqa: E712
        )
        .all()
    ):
        boq_by_uuid[str(it.id)] = it
        fac = fac_by_id.get(str(it.facility_id))
        if fac and it.original_code:
            boq_by_code[(fac.facility_code, str(it.original_code).strip())] = it

    # Auto-detect: ada sheet FAC_* → multi-sheet mode
    wb_check = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    fac_sheets = [s for s in wb_check.sheetnames if s.startswith("FAC_")]
    wb_check.close()

    parse_kwargs = dict(
        boq_by_uuid=boq_by_uuid,
        boq_by_code=boq_by_code,
        fac_index=fac_index,
        exclude_vo_id=exclude_vo_id,
        db=db,
    )

    all_items: List[Dict] = []
    all_warnings: List[str] = []
    all_errors: List[str] = []
    all_fac_codes: Set[str] = set()

    if fac_sheets:
        # Multi-sheet: parse setiap FAC_* sheet
        for sheet_name in fac_sheets:
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=0, dtype=object)
            items, warns, errs, fac_codes = _parse_df_rows(df, **parse_kwargs)
            all_items.extend(items)
            all_warnings.extend([f"[{sheet_name}] {w}" for w in warns])
            all_errors.extend([f"[{sheet_name}] {e}" for e in errs])
            all_fac_codes.update(fac_codes)
    else:
        # Flat: sheet pertama
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, header=0, dtype=object)
        items, warns, errs, fac_codes = _parse_df_rows(df, **parse_kwargs)
        all_items.extend(items)
        all_warnings.extend(warns)
        all_errors.extend(errs)
        all_fac_codes.update(fac_codes)

    return {
        "items": all_items,
        "warnings": all_warnings,
        "errors": all_errors,
        "facility_codes_in_file": sorted(all_fac_codes),
    }


def _safe_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    return str(v).strip()


def _safe_float(v) -> float:
    try:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return 0.0
        if isinstance(v, str) and not v.strip():
            return 0.0
        return float(v)
    except (TypeError, ValueError):
        return 0.0


_TWOPLACES = Decimal("0.01")


def _q2(v) -> Decimal:
    """Quantize to 2 decimal places (ROUND_HALF_UP). Aturan sistem: volume &
    harga satuan SELALU 2 dp, supaya vol_baru di Excel sama persis dengan
    volume aktif di DB tanpa float-noise.

    Defensive vs garbage input: None, NaN, Infinity, atau string non-numeric
    semua jadi Decimal('0.00') (Excel cell kosong sering masuk sebagai NaN
    dari pandas, dan Decimal.quantize meledak di NaN/Inf)."""
    from decimal import InvalidOperation
    if v is None:
        return Decimal("0.00")
    if isinstance(v, float) and (v != v or v in (float("inf"), float("-inf"))):
        return Decimal("0.00")
    if not isinstance(v, Decimal):
        try:
            v = Decimal(str(v))
        except (TypeError, ValueError, InvalidOperation):
            return Decimal("0.00")
    if v.is_nan() or v.is_infinite():
        return Decimal("0.00")
    return v.quantize(_TWOPLACES, rounding=ROUND_HALF_UP)


def _safe_decimal(v) -> Optional[Decimal]:
    """Convert Excel cell value to Decimal preserving the user-typed precision.
    Returns None when the cell is missing/empty/NaN/Inf (vs. 0 which is a
    real value). Goes via str() so float values use Python's shortest
    round-trip representation (e.g. 33.36 not 33.35999999999999943155847532...).
    """
    from decimal import InvalidOperation
    if v is None:
        return None
    if isinstance(v, float):
        if pd.isna(v) or v in (float("inf"), float("-inf")):
            return None
    if isinstance(v, str):
        v = v.strip()
        if not v:
            return None
    try:
        d = Decimal(str(v))
    except (TypeError, ValueError, InvalidOperation):
        return None
    if d.is_nan() or d.is_infinite():
        return None
    return d
