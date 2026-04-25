"""Generate Excel templates downloadable by users."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
NOTE_FILL = PatternFill("solid", fgColor="FEF3C7")


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _auto_width(ws, min_w=10, max_w=40):
    for col_cells in ws.columns:
        col = col_cells[0].column_letter
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col].width = min(max(length + 2, min_w), max_w)


# ─────────────────────────────────────────────────────────────────────────────

def template_boq_simple() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"

    headers = [
        "facility_code", "facility_name",
        "code", "parent_code",
        "description", "unit",
        "volume", "unit_price", "total_price",
        "planned_start_week", "planned_duration_weeks",
    ]
    ws.append(headers)
    _style_header(ws)

    examples = [
        ["GB-01", "Gudang Beku", "4",   "",    "PEKERJAAN PONDASI GUDANG BEKU", "",   0,     0,         0,          None, None],
        ["GB-01", "Gudang Beku", "A",   "4",   "PEKERJAAN STRUKTUR PONDASI",     "",   0,     0,         0,          None, None],
        ["GB-01", "Gudang Beku", "1",   "A",   "Pekerjaan Bouwplank",            "M",  42,    176668.75, 7420087.5,  1,    2],
        ["GB-01", "Gudang Beku", "2",   "A",   "Pekerjaan Pondasi Batu Belah",   "",   0,     0,         0,          None, None],
        ["GB-01", "Gudang Beku", "a",   "2",   "Penggalian tanah",               "M³", 33.36, 89929.71,  3000055.13, 1,    2],
    ]
    for ex in examples:
        ws.append(ex)

    # Instructions sheet
    ws2 = wb.create_sheet("Petunjuk")
    ws2["A1"] = "PETUNJUK PENGISIAN TEMPLATE BOQ"
    ws2["A1"].font = Font(bold=True, size=14)
    instructions = [
        "",
        "1. Isi sheet 'BOQ' sesuai kolom. Satu baris = satu item BOQ.",
        "2. facility_code: Kode unik fasilitas dalam lokasi (mis. GB-01 untuk Gudang Beku).",
        "   Semua baris dengan facility_code yang sama akan masuk ke fasilitas yang sama.",
        "3. facility_name: Nama fasilitas. Cukup diisi di baris pertama per facility_code.",
        "",
        "4. code: Kode lokal item dalam facility (mis. '4', 'A', '1', 'a').",
        "   - WAJIB UNIK per facility — dipakai sebagai referensi parent_code dari item lain.",
        "   - Boleh apa saja: angka, huruf, gabungan ('PER-006', 'STR.01').",
        "   - Kalau dikosongkan, sistem auto-generate (R1, R2, R3, …).",
        "",
        "5. parent_code: Kode item parent dalam facility yang sama.",
        "   - Kosong = item root (top-level, tidak di bawah siapa-siapa).",
        "   - Diisi = item ini menjadi child dari item ber-code = nilai ini.",
        "   - Inilah satu-satunya cara mengatur hirarki. Level auto-dihitung",
        "     dari rantai parent (root=0, anak root=1, cucu=2, dst).",
        "",
        "6. description: Uraian pekerjaan. Wajib diisi.",
        "7. volume, unit_price, total_price: hanya isi untuk leaf item (item paling bawah",
        "   di rantai, yang TIDAK punya child). Group/sub-group biarkan 0 atau kosong.",
        "   Kalau total_price kosong, dihitung dari volume × unit_price.",
        "8. planned_start_week, planned_duration_weeks: opsional, untuk schedule.",
        "9. Sistem otomatis hitung level, full_code (rantai code parent.code), is_leaf,",
        "   dan bobot %. Anda tidak perlu input itu manual.",
        "",
        "Tips: urutan baris di Excel tidak harus rapi parent-dulu-child. Selama",
        "parent_code merujuk ke code yang ada (di mana saja), sistem rangkai otomatis.",
    ]
    for i, line in enumerate(instructions, start=2):
        ws2[f"A{i}"] = line
    ws2.column_dimensions["A"].width = 90

    _auto_width(ws)
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def template_facilities() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Fasilitas"
    ws.append(["facility_code", "facility_name", "facility_type", "display_order", "notes"])
    _style_header(ws)
    sample = [
        ["GB-01", "Gudang Beku", "gudang_beku", 1, ""],
        ["PE-01", "Pabrik Es", "pabrik_es", 2, ""],
        ["KIOS-01", "Kios Perbekalan", "kios", 3, ""],
    ]
    for s in sample:
        ws.append(s)
    _auto_width(ws)
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def template_locations() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lokasi"
    ws.append(["location_code", "name", "village", "district", "city", "province", "latitude", "longitude"])
    _style_header(ws)
    ws.append(["LOK-01", "Desa Bintaro Ampean", "Bintaro", "Ampean", "Mataram", "NTB", "-8.5833", "116.1064"])
    _auto_width(ws)
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def template_weekly_progress(boq_items: list) -> bytes:
    """
    boq_items: list of dicts {id, full_code, description, unit, volume}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Progress"
    ws.append([
        "boq_item_id", "code", "description", "unit",
        "volume_boq", "volume_this_week", "volume_cumulative", "notes",
    ])
    _style_header(ws)
    for it in boq_items:
        ws.append([
            str(it["id"]),
            it.get("full_code") or it.get("original_code", ""),
            it["description"],
            it.get("unit") or "",
            float(it.get("volume") or 0),
            0,
            0,
            "",
        ])
    _auto_width(ws, max_w=60)
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()
